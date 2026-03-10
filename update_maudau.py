#!/usr/bin/env python3
import json
import os
import re
import shutil
import sys
import time
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import requests
from lxml import etree as ET

BASE_FEED_URL = "https://aqua-favorit.com.ua/content/export/b0026fd850ce11bb0cb7610e252d7dae.xml"
ROZETKA_FEED_URL = "http://parser.biz.ua/Aqua/api/export.aspx?action=rozetka&key=ui82P2VotQQamFTj512NQJK3HOlKvyv7"
MAUDAU_DIR = Path(__file__).resolve().parent / "maudau"
MAUDAU_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_XML = Path(__file__).resolve().parent / "update_maudau.xml"
LOCAL_OUTPUT_XML = MAUDAU_DIR / "update_maudau.xml"
MAPPING_REPORT_XLSX = MAUDAU_DIR / "maudau_mapping_gaps.xlsx"
FINAL_MAPPING_TEMPLATE_XLSX = MAUDAU_DIR / "Финал_мапинг.xlsx"
FINAL_MAPPING_REPORT_XLSX = MAUDAU_DIR / "Попытка1_готово.xlsx"
CATEGORY_LIST_CANDIDATES = [
    Path("/Volumes/X-Files/Загрузки рабочие/Maudau/2/categories-2026-02-18-0752.xlsx"),
    MAUDAU_DIR / "categories-2026-02-18-0752.xlsx",
    Path("categories-2026-02-18-0752.xlsx"),
]
BRANDS_LIST_CANDIDATES = [
    Path("/Volumes/X-Files/Загрузки рабочие/brands-2026-03-09-2147.xlsx"),
    MAUDAU_DIR / "brands-2026-03-09-2147.xlsx",
    Path("brands-2026-03-09-2147.xlsx"),
]
COUNTRIES_LIST_CANDIDATES = [
    Path("/Volumes/X-Files/Загрузки рабочие/countries-2026-03-10-1651.xlsx"),
    MAUDAU_DIR / "countries-2026-03-10-1651.xlsx",
    Path("countries-2026-03-10-1651.xlsx"),
]

TMP_DIR = Path("/tmp/maudau_feed")
TMP_DIR.mkdir(parents=True, exist_ok=True)
BASE_XML = TMP_DIR / "base.xml"
ROZETKA_XML = TMP_DIR / "rozetka.xml"
BACKUP_DIR = MAUDAU_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
BASE_BACKUP_XML = BACKUP_DIR / "aquafavorit_last.xml"
ROZETKA_BACKUP_XML = BACKUP_DIR / "parserbiz_last.xml"
SOURCES_STATE_JSON = BACKUP_DIR / "sources_state.json"
SOURCE_STALE_HOURS = 72
LOCAL_ENV_FILE = Path(__file__).resolve().parent / ".env"
ROZETKA_BACKUP_CANDIDATES = [
    Path(os.environ.get("ROZETKA_LOCAL_XML", "")).expanduser(),
    ROZETKA_BACKUP_XML,
    MAUDAU_DIR / "https-::parser.biz.ua:Aqua:api:export.aspx?action=rozetka&key=ui82P2VotQQamFTj512NQJK3HOlKvyv7",
    MAUDAU_DIR / "rozetka.xml",
    Path("rozetka.xml"),
    ROZETKA_XML,
]

BASE_BACKUP_CANDIDATES = [
    Path(os.environ.get("BASE_LOCAL_XML", "")).expanduser(),
    BASE_BACKUP_XML,
    MAUDAU_DIR / "aquafavorit.xml",
    MAUDAU_DIR / "base.xml",
    Path("aquafavorit.xml"),
    BASE_XML,
]

MERCHANT_CATEGORIES_CANDIDATES = [
    Path(os.environ.get("MAUDAU_MERCHANT_CATEGORIES_XML", "")).expanduser(),
    MAUDAU_DIR / "merchant_categories_2026-02-18-0837.xml",
    Path("merchant_categories_2026-02-18-0837.xml"),
    Path("/Volumes/X-Files/Загрузки рабочие/Maudau/2/merchant_categories_2026-02-18-0837.xml"),
    Path("/Volumes/X-Files/Загрузки рабочие/Maudau/merchant_categories_2026-02-16-1420.xml"),
    Path("merchant_categories_2026-02-16-1420.xml"),
    Path("/Users/Kezhunya/Downloads/merchant_categories_2026-03-09-1229.xml"),
    MAUDAU_DIR / "merchant_categories_2026-03-09-1229.xml",
    Path("merchant_categories_2026-03-09-1229.xml"),
]

# source category id -> maudau category id
SOURCE_TO_MAUDAU_CATEGORY = {
    # Baths
    "1059": "3054",
    "1060": "3054",
    "1061": "3054",
    "1062": "3054",
    "1150": "3054",
    # Bath accessories
    "1143": "1412",
    "1139": "1417",
    "1140": "1418",
    "1137": "1421",
    "1265": "1423",
    "1136": "1424",
    "1144": "1427",
    "1066": "1428",
    "1067": "3172",
    "1138": "1429",
    "1264": "1429",
    "1145": "1430",
    "1190": "1431",
    "1173": "2313",
    # Mixers / showers
    "1007": "1899",
    "1068": "1899",
    "1069": "1899",
    "1070": "1899",
    "1071": "1899",
    "1072": "1899",
    "1073": "1899",
    "1211": "1899",
    "1214": "1899",
    "1217": "1899",
    "1242": "1899",
    "1224": "1899",
    "1075": "2299",
    "1100": "1904",
    "1101": "1904",
    "1099": "2214",
    "1103": "2214",
    "1104": "2223",
    "1105": "2212",
    "1106": "2223",
    "1109": "2223",
    "1110": "1906",
    "1111": "2957",
    "1118": "3172",
    # Sanitary ware
    "1088": "2393",
    "1080": "1908",
    "1081": "1908",
    "1082": "1908",
    "1083": "1908",
    "1084": "1908",
    "1085": "1907",
    "1089": "2393",
    # Basins
    "1094": "3049",
    "1095": "3049",
    "1096": "3049",
    "1097": "3049",
    "1098": "3049",
    # Bathroom furniture / kitchen sinks
    "1129": "2484",
    "1131": "2920",
    "1132": "2920",
    "1176": "2483",
    "1166": "1903",
    "1168": "3165",
    "1141": "1175",
    "1171": "1411",
    "1189": "1411",
    "1206": "1024",
    # Heating / boilers / ventilation
    "1158": "2398",
    "1160": "1897",
    "1090": "3172",
    "1167": "3189",
    # Heaters (from дополнение.xlsx)
    "1228": "669",
    "1266": "671",
    "1227": "669",
    "1226": "669",
    # Engineering
    "1253": "1740",
    # Toilets special
    "1175": "1907",
    # Towels / radiators
    "1215": "1901",
    "1119": "1902",
    "1120": "1902",
    # Shower (per manual mapping)
    "1102": "1904",
}

# Fallback names for Maudau category ids that may be absent in primary merchant_categories dump.
MAUDAU_CATEGORY_NAME_OVERRIDES = {
    "3175": "Электрические ТЭНы для сушилок для полотенец",
    "3189": "Аксессуары к сушилке для полотенец и радиаторов",
    "3172": "Сифоны",
}

COUNTRY_ALIASES_RU_TO_UA = {
    "германия": "Німеччина",
    "польша": "Польща",
    "украина": "Україна",
    "чехия": "Чехія",
    "испания": "Іспанія",
    "турция": "Туреччина",
    "италия": "Італія",
    "швейцария": "Швейцарія",
    "индия": "Індія",
    "австрия": "Австрія",
    "венгрия": "Угорщина",
    "дания": "Данія",
    "болгария": "Болгарія",
    "словения": "Словенія",
    "португалия": "Португалія",
    "великобритания": "Великобританія",
    "великобританія": "Великобританія",
}

# Categories explicitly marked as "will be created on Maudau later (no edits now)".
SKIP_REMAP_SOURCE_CATEGORIES = {
    "1064",
    "1076",
    "1086",
    "1107",
    "1108",
    "1112",
    "1113",
    "1114",
    "1116",
    "1122",
    "1123",
    "1124",
    "1125",
    "1127",
    "1130",
    "1133",
    "1134",
    "1153",
    "1161",
    "1201",
    "1202",
    "1205",
    "1219",
    "1223",
    "1232",
    "1250",
    "1255",
    "1257",
    "1271",
    "1272",
}

# Categories with unresolved target ("?") in final mapping template.
QUESTION_SOURCE_CATEGORIES = {
    "1121",
    "1126",
    "1156",
    "1249",
    "1256",
}

# Keep these source categories in feed even if offer is absent in Rozetka.
# This extends the vendor-level exception and is applied before Rozetka pruning.
KEEP_WITHOUT_ROZETKA_SOURCE_CATEGORIES = {
    # Heaters
    "1228",  # Керамические обогреватели
    "1227",  # Металлокерамические обогреватели
    "1226",  # Теплый плинтус
    "1266",  # Инфрокрасные/инфракрасные обогреватели
    # Warm floor
    "1271",  # Нагревательные маты
    "1272",  # Нагревательные кабели
}

# Always include these source categories in the XLSX layout blocks,
# even when current source XML has zero offers for them.
FORCE_LAYOUT_SOURCE_CATEGORIES = {
    "1228",  # Керамические обогреватели
    "1227",  # Металлокерамические обогреватели
    "1226",  # Теплый плинтус
    "1266",  # Инфракрасные обогреватели
}

# Source towel category id -> forced Type value for Maudau category 1902
TOWEL_TYPE_BY_SOURCE_CATEGORY = {
    "1119": "Водяные",
    "1120": "Электрические",
}

# Source category id -> list of forced Maudau attributes/values.
# Used when several source subcategories collapse into one Maudau category.
FORCED_ATTRS_BY_SOURCE_CATEGORY = {
    # Mixers
    "1068": [("Тип", "Наборы смесителей")],
    "1069": [("Тип", "Смесители"), ("Застосування", "Для умывальника")],
    "1070": [("Тип", "Смесители"), ("Застосування", "Для ванной")],
    "1071": [("Тип", "Смесители"), ("Застосування", "Для биде")],
    "1072": [("Тип", "Смесители"), ("Застосування", "Для душа")],
    "1073": [("Тип", "Смесители"), ("Застосування", "Для кухни")],
    "1214": [("Тип", "Смесители"), ("Вид", "Одновентильные")],
    "1224": [("Тип", "Смесители"), ("Застосування", "Для кухни")],
    # Shower rods/holders
    "1104": [("Тип", "Душевая штанга")],
    "1106": [("Тип", "Держатель душа")],
    "1099": [("Вид лійки", "Верхний душ")],
    "1103": [("Вид лійки", "Ручной душ")],
    # Bathroom dispensers
    "1138": [("Призначення", "Для мыла")],
    "1264": [("Призначення", "Для антисептика")],
}

# Common RU source param name -> Maudau attribute nameUK
COMMON_PARAM_NAME_MAP = {
    "материал": "Матеріал",
    "цвет": "Колір",
    "высота, см": "Висота",
    "высота, мм": "Висота",
    "ширина, см": "Ширина",
    "ширина, мм": "Ширина",
    "длинна, см": "Довжина",
    "длина, см": "Довжина",
    "длина, мм": "Довжина",
    "глубина, см": "Глибина",
    "глубина, мм": "Глибина",
    "форма": "Форма",
    "тип": "Тип",
    "назначение": "Призначення",
    "установка": "Встановлення",
    "монтаж": "Монтаж",
    "способ монтажа": "Тип монтажу",
    "тип установки": "Тип установки",
    "оснащение": "Оснащення",
    "количество чаш": "Кількість чаш",
    "подключение": "Підключення",
    "тип подключения": "Підключення води",
    "диаметр подключения": "Діаметр підключення",
    "вес, кг": "Вага",
    "гарантия": "Гарантія",
    "гарантия, мес.": "Гарантія",
    "гарантийный срок, мес.": "Гарантія",
    "гарантийный срок, мес": "Гарантія",
    "поддон": "Піддон",
    "тип дверей": "Тип відчинення дверей",
    "длина душевого шланга, см": "Довжина душового шлангу",
    "длина шланга, см": "Довжина шлангу",
    "диаметр душа, мм": "Діаметр ручної лійки",
    "количество режимов": "Кількість режимів",
    "комплектация": "Комплектація",
    "особенности": "Особливості",
}

# Category-specific overrides: target maudau category -> source param key -> maudau attr nameUK
CATEGORY_PARAM_NAME_OVERRIDES = {
    "1899": {
        "управление": "Вид",
    },
    "1901": {
        "тип установки": "Монтаж",
        "вид радиатора": "Вид",
        "мощность, вт": "Теплова потужність",
    },
    "1902": {
        "тип подключения": "Підключення води",
        "форма": "Вид",
        "мощность, вт": "Максимальна потужність",
    },
    "1904": {
        "количество режимов": "Кількість режимів струменя ручного душу",
        "длина душевого шланга, см": "Довжина душового шлангу",
        "диаметр душа, мм": "Діаметр ручної лійки",
    },
    "2214": {
        "количество режимов": "Кількість режимів",
        "форма лейки": "Форма",
    },
    "2223": {
        "тип установки": "Тип монтажу",
        "назначение": "Призначення",
    },
    "2299": {
        "количество режимов": "Кількість режимів струменя ручної лійки",
        "диаметр душа, мм": "Діаметр ручної лійки",
        "тип": "Вид змішувача",
        "установка": "Встановлення",
    },
    "669": {
        "тип установки": "Установка",
        "мощность, вт": "Потужність",
        "площадь обслуживания, кв. м": "Площа обігріву",
        "управление": "Управління",
        "оснащение": "Особливості",
        "цвет": "Колір",
    },
    "671": {
        "тип установки": "Установка",
        "мощность, вт": "Потужність",
        "площадь обслуживания, кв. м": "Площа обігріву",
        "управление": "Управління",
        "оснащение": "Особливості",
        "цвет": "Колір",
    },
}

# Category value normalization: target category -> attr -> source value -> canonical Maudau value.
CATEGORY_ATTR_VALUE_OVERRIDES = {
    "1899": {
        "Вид": {
            "рычаг": "Однорычажные",
            "однорычажный": "Однорычажные",
            "однорычажные": "Однорычажные",
            "вентиль": "Двухвентильные",
            "вентильный": "Двухвентильные",
            "двухвентильный": "Двухвентильные",
            "сенсор": "Сенсорные",
            "сенсорный": "Сенсорные",
            "джойстик": "Джойстиковые",
            "джойстиковый": "Джойстиковые",
            "кнопочный": "Нажимные",
            "порционный": "Нажимные",
            "рукоятка select": "Однорычажные",
        },
    },
    "669": {
        "Установка": {
            "настенный": "Настенные",
            "настенная": "Настенные",
            "горизонтальный": "Настенные",
            "вертикальный": "Настенные",
            "горизонтальный/вертикальный": "Настенные",
            "напольный": "Напольные",
            "напольная": "Напольные",
            "настольный": "Настольные",
            "потолочный": "Потолочные",
        },
        "Управління": {
            "регулятор": "Механическое",
            "механическое": "Механическое",
            "электронное": "Электронное",
        },
        "Особливості": {
            "регулятор температуры|шнур с вилкой": "С терморегулятором",
            "регулятор температуры": "С терморегулятором",
        },
    },
    "671": {
        "Установка": {
            "настенный": "Настенные",
            "настенная": "Настенные",
            "горизонтальный": "Настенные",
            "вертикальный": "Настенные",
            "горизонтальный/вертикальный": "Настенные",
            "напольный": "Напольные",
            "напольная": "Напольные",
            "настольный": "Настольные",
            "потолочный": "Потолочные",
        },
        "Управління": {
            "регулятор": "Механическое",
            "механическое": "Механическое",
            "электронное": "Электронное",
            "сенсорное": "Сенсорное",
        },
    },
}

ALLOWED_VENDORS = {"мойдодыр", "dusel"}
OLD_PRICE_TAGS = ("old_price", "oldprice", "price_old", "old", "priceold")
ID_CLEAN_RE = re.compile(r"[^A-Za-z0-9]")
HTML_TAG_RE = re.compile(r"<[^>]+>")
CYRILLIC_RE = re.compile(r"[\u0400-\u04FF]")
MULTISPACE_RE = re.compile(r"\s+")
VALUE_UNIT_RE = re.compile(r"^\s*(\d+(?:[\.,]\d+)?)\s*(мм|см|м|вт|квт|квт\.|мес\.?|месяц(?:а|ев)?|год(?:а|ов)?)\s*$", re.IGNORECASE)
RANGE_SIMPLE_RE = re.compile(r"^\s*(\d+(?:[.,]\d+)?)\s*[-–]\s*(\d+(?:[.,]\d+)?)\s*([^\d]*)$", re.IGNORECASE)
RANGE_UPPER_RE = re.compile(r"^\s*(до)\s*(\d+(?:[.,]\d+)?)\s*([^\d]*)$", re.IGNORECASE)
RANGE_LOWER_RE = re.compile(r"^\s*(более|понад|більше)\s*(\d+(?:[.,]\d+)?)\s*([^\d]*)$", re.IGNORECASE)
SCALAR_WITH_UNIT_RE = re.compile(r"^\s*(\d+(?:[.,]\d+)?)\s*([^\d]*)$", re.IGNORECASE)
MAX_DESC_LEN = 400

# Generic value normalization hints (RU/UA forms and common endings).
GENERIC_VALUE_SYNONYMS = {
    "отдельностоящая": "Отдельно стоящие",
    "отдельностоящий": "Отдельно стоящие",
    "отдельно стоящая": "Отдельно стоящие",
    "пристенная": "Пристенные",
    "встроенная": "Встроенные",
    "прямоугольная": "Прямоугольные",
    "прямокутна": "Прямокутні",
    "круглая": "Круглые",
    "кругла": "Круглі",
    "квадратная": "Квадратные",
    "квадратна": "Квадратні",
    "овальная": "Овальные",
    "овальна": "Овальні",
    "асимметричная": "Асимметричные",
    "асиметрична": "Асиметричні",
    "левый": "Асимметричные левые",
    "правый": "Асимметричные правые",
    "белая": "Белый",
    "черная": "Черный",
}

def load_local_env(path: Path) -> None:
    if not path.exists():
        return
    try:
        for raw_line in path.read_text(encoding="utf-8").splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            if line.startswith("export "):
                line = line[7:].strip()
            key, value = line.split("=", 1)
            key = key.strip()
            value = value.strip()
            if not key or key in os.environ:
                continue
            if len(value) >= 2 and value[0] == value[-1] and value[0] in {"'", '"'}:
                value = value[1:-1]
            os.environ[key] = value
    except Exception as exc:
        print(f"⚠ Не удалось прочитать .env ({path}): {exc}")


load_local_env(LOCAL_ENV_FILE)
TG_BOT_TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN")
TG_CHAT_ID = os.environ.get("TELEGRAM_CHAT_ID")


def send_telegram(message: str) -> None:
    if not TG_BOT_TOKEN or not TG_CHAT_ID:
        print("⚠ TELEGRAM_BOT_TOKEN или TELEGRAM_CHAT_ID не задан. Сообщение не отправлено.")
        return
    url = f"https://api.telegram.org/bot{TG_BOT_TOKEN}/sendMessage"
    payload = {"chat_id": TG_CHAT_ID, "text": message}
    try:
        resp = requests.post(url, data=payload, timeout=20)
        resp.raise_for_status()
    except Exception as exc:
        print(f"⚠ Ошибка отправки в Telegram: {exc}")


def download_file(url: str, path: Path, title: str, retries: int = 5, timeout: int = 180) -> None:
    print(f"▶ Загрузка: {title}")
    for attempt in range(1, retries + 1):
        try:
            r = requests.get(url, stream=True, timeout=timeout)
            r.raise_for_status()
            with open(path, "wb") as f:
                for chunk in r.iter_content(1024 * 1024):
                    if chunk:
                        f.write(chunk)
            print(f"✅ {title} загружен")
            return
        except Exception as exc:
            print(f"⚠ Ошибка загрузки ({title}) попытка {attempt}/{retries}: {exc}")
            if attempt == retries:
                raise
            time.sleep(5)


def resolve_rozetka_backup_path() -> Path | None:
    for candidate in ROZETKA_BACKUP_CANDIDATES:
        if not str(candidate):
            continue
        if candidate.is_file() and candidate.stat().st_size > 0:
            try:
                ET.parse(str(candidate))
            except Exception:
                continue
            return candidate
    return None


def resolve_base_backup_path() -> Path | None:
    for candidate in BASE_BACKUP_CANDIDATES:
        if not str(candidate):
            continue
        if candidate.is_file() and candidate.stat().st_size > 0:
            try:
                ET.parse(str(candidate))
            except Exception:
                continue
            return candidate
    return None


def load_sources_state() -> dict:
    if not SOURCES_STATE_JSON.exists():
        return {}
    try:
        return json.loads(SOURCES_STATE_JSON.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_sources_state(state: dict) -> None:
    try:
        SOURCES_STATE_JSON.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception as exc:
        print(f"⚠ Не удалось сохранить sources_state.json: {exc}")


def update_source_success(state: dict, source_key: str, used_path: Path) -> None:
    state[source_key] = {
        "last_success_ts": int(time.time()),
        "last_success_path": str(used_path),
        "first_failure_ts": None,
        "last_failure_ts": None,
    }


def update_source_failure(state: dict, source_key: str) -> None:
    now_ts = int(time.time())
    item = state.get(source_key, {})
    first_failure = item.get("first_failure_ts")
    if not isinstance(first_failure, int):
        first_failure = now_ts
    item["first_failure_ts"] = first_failure
    item["last_failure_ts"] = now_ts
    state[source_key] = item


def stale_alert_text(state: dict, source_key: str, source_label: str, backup_path: Path | None = None) -> str:
    item = state.get(source_key, {})
    ref_ts: int | None = None

    last_success_ts = item.get("last_success_ts")
    if isinstance(last_success_ts, int):
        ref_ts = last_success_ts
    elif backup_path and backup_path.exists():
        try:
            ref_ts = int(backup_path.stat().st_mtime)
        except Exception:
            ref_ts = None
    else:
        first_failure_ts = item.get("first_failure_ts")
        if isinstance(first_failure_ts, int):
            ref_ts = first_failure_ts

    if ref_ts is None:
        return ""

    hours = (time.time() - ref_ts) / 3600.0
    if hours >= SOURCE_STALE_HOURS:
        return f"{source_label} недоступен 72 часа"
    return ""


def normalize_text(value: str | None) -> str:
    return (value or "").strip()


def normalize_key(value: str | None) -> str:
    return normalize_text(value).casefold()


def build_source_category_names(root: ET._Element) -> dict[str, str]:
    items: dict[str, str] = {}
    parent_by_id: dict[str, str] = {}

    for c in root.xpath("//shop/categories/category"):
        cid = normalize_text(c.get("id"))
        if not cid:
            continue
        items[cid] = normalize_text(c.text)
        parent = normalize_text(c.get("parentId"))
        if parent:
            parent_by_id[cid] = parent

    def is_generic(name: str) -> bool:
        key = normalize_key(name)
        return ("комплектующ" in key) or ("аксессуар" in key)

    def resolve_parent_section(cid: str) -> str:
        cur = parent_by_id.get(cid, "")
        seen: set[str] = set()
        while cur and cur not in seen:
            seen.add(cur)
            candidate = items.get(cur, "")
            if candidate and not is_generic(candidate):
                return candidate
            cur = parent_by_id.get(cur, "")
        return ""

    result: dict[str, str] = {}
    for cid, name in items.items():
        display = name
        if is_generic(name):
            section = resolve_parent_section(cid)
            if section:
                display = f"{name} ({section})"
        result[cid] = display

    # If the same visible name appears in several categories, disambiguate by parent section.
    count_by_name: dict[str, int] = defaultdict(int)
    for name in result.values():
        count_by_name[name] += 1

    for cid, display in list(result.items()):
        if count_by_name.get(display, 0) <= 1:
            continue
        if "(" in display and ")" in display:
            continue
        section = resolve_parent_section(cid)
        if section:
            result[cid] = f"{display} ({section})"

    return result


def child_text(offer: ET._Element, tag: str) -> str:
    node = offer.find(tag)
    return normalize_text(node.text if node is not None else "")


def compact_text(value: str) -> str:
    return normalize_text(MULTISPACE_RE.sub(" ", value))


def find_param_value(offer: ET._Element, param_name: str) -> str:
    target = normalize_key(param_name)
    for param in offer.findall("param"):
        if normalize_key(param.get("name")) == target:
            value = normalize_text(param.text)
            if value:
                return value
    return ""


def upsert_param(offer: ET._Element, param_name: str, value: str) -> bool:
    clean_name = normalize_text(param_name)
    clean_value = normalize_text(value)
    if not clean_name or not clean_value:
        return False

    for param in offer.findall("param"):
        if normalize_key(param.get("name")) == normalize_key(clean_name):
            if normalize_text(param.text) != clean_value:
                param.text = clean_value
                param.set("name", clean_name)
                return True
            return False

    p = ET.SubElement(offer, "param")
    p.set("name", clean_name)
    p.text = clean_value
    return True


def resolve_offer_id_raw(offer: ET._Element) -> str:
    article = find_param_value(offer, "Артикул")
    if article:
        return article
    vendor_code = child_text(offer, "vendorCode")
    if vendor_code:
        return vendor_code
    return normalize_text(offer.get("id"))


def resolve_offer_id_key(offer: ET._Element) -> str:
    return normalize_key(resolve_offer_id_raw(offer))


def extract_old_price(offer: ET._Element) -> str:
    for tag in OLD_PRICE_TAGS:
        value = child_text(offer, tag)
        if value:
            return value
    return ""


def extract_available(offer: ET._Element) -> str:
    return normalize_text(offer.get("available"))


def set_or_create(offer: ET._Element, tag: str, value: str) -> bool:
    value = normalize_text(value)
    if not value:
        return False
    node = offer.find(tag)
    if node is None:
        node = ET.SubElement(offer, tag)
        node.text = value
        return True
    if normalize_text(node.text) != value:
        node.text = value
        return True
    return False


def set_available(offer: ET._Element, value: str) -> bool:
    value = normalize_text(value)
    if value not in {"true", "false"}:
        value = "false"
    if normalize_text(offer.get("available")) == value:
        return False
    offer.set("available", value)
    return True


def normalize_name_description(offer: ET._Element) -> None:
    name = offer.find("name")
    name_ru = offer.find("name_ru")
    name_ua = offer.find("name_ua")

    if name is not None and name_ru is None:
        name_ru = ET.SubElement(offer, "name_ru")
        name_ru.text = normalize_text(name.text)
    if name is not None:
        offer.remove(name)

    if name_ru is not None:
        name_ru.text = normalize_text(HTML_TAG_RE.sub("", name_ru.text or ""))[:255]
    if name_ua is not None:
        name_ua.text = normalize_text(HTML_TAG_RE.sub("", name_ua.text or ""))[:255]

    desc = offer.find("description")
    desc_ru = offer.find("description_ru")
    desc_ua = offer.find("description_ua")

    if desc is not None and desc_ru is None:
        desc_ru = ET.SubElement(offer, "description_ru")
        desc_ru.text = normalize_text(desc.text)
    if desc is not None:
        offer.remove(desc)

    if desc_ru is not None:
        desc_ru.text = compact_text(HTML_TAG_RE.sub(" ", desc_ru.text or ""))[:MAX_DESC_LEN]
    if desc_ua is not None:
        desc_ua.text = compact_text(HTML_TAG_RE.sub(" ", desc_ua.text or ""))[:MAX_DESC_LEN]


def normalize_old_price(offer: ET._Element) -> None:
    values = {}
    for child in list(offer):
        tag = child.tag
        if tag in OLD_PRICE_TAGS:
            value = normalize_text(child.text)
            if value and tag not in values:
                values[tag] = value
            offer.remove(child)

    old_value = ""
    for tag in OLD_PRICE_TAGS:
        if values.get(tag):
            old_value = values[tag]
            break

    if old_value:
        node = ET.SubElement(offer, "old_price")
        node.text = old_value


def enrich_vendor_country_from_params(offer: ET._Element) -> int:
    changed = 0
    brand = (
        find_param_value(offer, "Бренд")
        or find_param_value(offer, "Бренд товара")
        or find_param_value(offer, "Торговая марка")
    )
    reg_country = (
        find_param_value(offer, "Страна регистрации бренда")
        or find_param_value(offer, "Країна реєстрації бренду")
    )

    if brand:
        if set_or_create(offer, "vendor", compact_text(brand)):
            changed += 1
    if reg_country:
        # Requested mapping: Країна-виробник = Страна регистрации бренда.
        if set_or_create(offer, "country", compact_text(reg_country)):
            changed += 1

    return changed


def normalize_vendor_by_catalog(offer: ET._Element, brands_catalog: dict[str, str]) -> bool:
    if not brands_catalog:
        return False
    vendor = child_text(offer, "vendor")
    if not vendor:
        return False
    canonical = brands_catalog.get(normalize_text_key(vendor))
    if not canonical:
        return False
    return set_or_create(offer, "vendor", canonical)


def normalize_country_by_catalog(offer: ET._Element, countries_catalog: dict[str, str]) -> bool:
    if not countries_catalog:
        return False
    country = child_text(offer, "country")
    if not country:
        return False

    key = normalize_text_key(country)
    canonical = countries_catalog.get(key)
    if not canonical:
        alias = COUNTRY_ALIASES_RU_TO_UA.get(normalize_key(country))
        if alias:
            canonical = countries_catalog.get(normalize_text_key(alias), alias)
    if not canonical:
        return False
    return set_or_create(offer, "country", canonical)


def resolve_merchant_categories_paths() -> list[Path]:
    paths: list[Path] = []
    seen: set[str] = set()
    for candidate in MERCHANT_CATEGORIES_CANDIDATES:
        if not str(candidate) or not candidate.is_file():
            continue
        resolved = str(candidate.resolve())
        if resolved in seen:
            continue
        seen.add(resolved)
        paths.append(candidate)
    return paths


def resolve_merchant_categories_path() -> Path | None:
    paths = resolve_merchant_categories_paths()
    return paths[0] if paths else None


def load_merchant_catalog(path: Path | None) -> dict[str, dict]:
    if path is None:
        return {}

    tree = ET.parse(str(path))
    root = tree.getroot()
    catalog: dict[str, dict] = {}

    for category in root.xpath(".//category[@portal_id]"):
        cid = normalize_text(category.get("portal_id"))
        if not cid:
            continue

        attrs: dict[str, dict[str, str]] = {}
        attr_lookup: dict[str, str] = {}

        for attr in category.findall("attribute"):
            attr_name = normalize_text(attr.get("nameUK"))
            if not attr_name:
                continue

            values_map = attrs.setdefault(attr_name, {})
            attr_lookup[normalize_text_key(attr_name)] = attr_name

            for value in attr.findall("attribute_value"):
                ru = normalize_text(value.get("nameRU"))
                uk = normalize_text(value.get("nameUK"))
                if ru:
                    values_map[normalize_text_key(ru)] = ru
                elif uk:
                    values_map[normalize_text_key(uk)] = uk

        catalog[cid] = {
            "name_ru": normalize_text(category.get("nameRU")),
            "name_uk": normalize_text(category.get("nameUK")),
            "attrs": attrs,
            "attr_lookup": attr_lookup,
        }

    return catalog


def merge_merchant_catalogs(base: dict[str, dict], extra: dict[str, dict]) -> dict[str, dict]:
    merged = dict(base)
    for cid, meta in extra.items():
        if cid not in merged:
            merged[cid] = meta
            continue
        target = merged[cid]
        if not target.get("name_ru") and meta.get("name_ru"):
            target["name_ru"] = meta.get("name_ru")
        if not target.get("name_uk") and meta.get("name_uk"):
            target["name_uk"] = meta.get("name_uk")

        target_lookup = target.setdefault("attr_lookup", {})
        target_attrs = target.setdefault("attrs", {})
        for attr_name, values_map in meta.get("attrs", {}).items():
            target_attrs.setdefault(attr_name, {})
            target_attrs[attr_name].update(values_map)
        target_lookup.update(meta.get("attr_lookup", {}))
    return merged


def resolve_category_list_path() -> Path | None:
    for candidate in CATEGORY_LIST_CANDIDATES:
        if candidate.is_file():
            return candidate
    return None


def load_category_names_from_xlsx(path: Path | None) -> dict[str, str]:
    if path is None:
        return {}
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}

    wb = load_workbook(str(path), read_only=True)
    sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    names: dict[str, str] = {}
    # Header expected: ID ... Назва (російською) at col 4.
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        raw_id = row[0]
        raw_name_ru = row[3] if len(row) > 3 else None
        if raw_id is None:
            continue
        cid = str(int(raw_id)) if isinstance(raw_id, (int, float)) else normalize_text(str(raw_id))
        name_ru = normalize_text(str(raw_name_ru)) if raw_name_ru is not None else ""
        if cid and name_ru:
            names[cid] = name_ru
    return names


def resolve_brand_list_path() -> Path | None:
    for candidate in BRANDS_LIST_CANDIDATES:
        if candidate.is_file():
            return candidate
    return None


def load_brands_catalog(path: Path | None) -> dict[str, str]:
    if path is None:
        return {}
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}

    wb = load_workbook(str(path), read_only=True)
    sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    result: dict[str, str] = {}
    for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        if not row:
            continue
        raw = row[0]
        if raw is None:
            continue
        brand = normalize_text(str(raw))
        if not brand:
            continue
        # Skip common header row(s).
        if idx == 1 and normalize_key(brand) in {"назва бренду", "бренд", "brand"}:
            continue
        result.setdefault(normalize_text_key(brand), brand)
    return result


def resolve_countries_list_path() -> Path | None:
    for candidate in COUNTRIES_LIST_CANDIDATES:
        if candidate.is_file():
            return candidate
    return None


def load_countries_catalog(path: Path | None) -> dict[str, str]:
    if path is None:
        return {}
    try:
        from openpyxl import load_workbook
    except Exception:
        return {}

    wb = load_workbook(str(path), read_only=True)
    sheet = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]

    result: dict[str, str] = {}
    for idx, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        if not row:
            continue
        raw = row[0]
        if raw is None:
            continue
        country = normalize_text(str(raw))
        if not country:
            continue
        if idx == 1 and normalize_key(country) in {"назва країни", "країна", "country"}:
            continue
        result.setdefault(normalize_text_key(country), country)
    return result


def normalize_number(num: float) -> str:
    if abs(num - round(num)) < 1e-9:
        return str(int(round(num)))
    text = f"{num:.3f}".rstrip("0").rstrip(".")
    return text


def unit_value_candidates(value: str) -> set[str]:
    raw = compact_text(value)
    candidates = {raw}

    match = VALUE_UNIT_RE.match(raw)
    if not match:
        return candidates

    num_text = match.group(1).replace(",", ".")
    unit = normalize_key(match.group(2)).replace(".", "")

    try:
        num = float(num_text)
    except ValueError:
        return candidates

    def add(val: float, unit_text: str) -> None:
        candidates.add(f"{normalize_number(val)} {unit_text}")

    if unit == "мм":
        add(num / 10, "см")
    elif unit == "см":
        add(num * 10, "мм")
    elif unit == "вт":
        add(num / 1000, "кВт")
    elif unit in {"квт"}:
        add(num * 1000, "Вт")
    elif unit in {"мес", "месяца", "месяцев"}:
        add(num, "мес")
        if abs(num % 12) < 1e-9:
            years = num / 12
            add(years, "год")
            add(years, "года")
    elif unit in {"год", "года", "годов"}:
        add(num, "год")
        add(num * 12, "мес")

    return candidates


def map_param_name(param_name: str, target_category_id: str) -> str:
    key = normalize_key(param_name)
    overrides = CATEGORY_PARAM_NAME_OVERRIDES.get(target_category_id, {})
    if key in overrides:
        return overrides[key]
    return COMMON_PARAM_NAME_MAP.get(key, normalize_text(param_name))


def pick_allowed_value(
    merchant_catalog: dict[str, dict],
    target_category_id: str,
    attr_name: str,
    candidates: list[str],
) -> str:
    meta = merchant_catalog.get(target_category_id, {})
    attrs = meta.get("attrs", {})
    values = attrs.get(attr_name, {})
    if not values:
        return ""

    for candidate in candidates:
        if not candidate:
            continue
        key = normalize_key(candidate)
        if key in values:
            return values[key]
    return ""


def infer_forced_attrs_from_source_name(
    source_name: str,
    target_category_id: str,
    merchant_catalog: dict[str, dict],
) -> list[tuple[str, str]]:
    name_key = normalize_key(source_name)
    inferred: list[tuple[str, str]] = []

    if target_category_id == "1899":
        app_map = {
            "умывальника": "Для умывальника",
            "кухни": "Для кухни",
            "ванны": "Для ванной",
            "душа": "Для душа",
            "биде": "Для биде",
            "гигиеничес": "Для гигиенического душа",
            "писсуар": "Для писсуаров",
            "мойки": "Для мойки",
        }
        for key_part, app_value in app_map.items():
            if key_part in name_key:
                picked = pick_allowed_value(
                    merchant_catalog,
                    target_category_id,
                    "Застосування",
                    [app_value],
                )
                if picked:
                    inferred.append(("Застосування", picked))
                break

        if "комплект" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Тип",
                ["Наборы смесителей"],
            )
            if picked:
                inferred.append(("Тип", picked))
        elif "смесител" in name_key or "кран" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Тип",
                ["Смесители"],
            )
            if picked:
                inferred.append(("Тип", picked))

        if "монокран" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Вид",
                ["Одновентильные"],
            )
            if picked:
                inferred.append(("Вид", picked))
        if "дозатор" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Вид",
                ["Нажимные"],
            )
            if picked:
                inferred.append(("Вид", picked))

    if target_category_id == "2223":
        if "штанг" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Тип",
                ["Душевая штанга"],
            )
            if picked:
                inferred.append(("Тип", picked))
        elif "держател" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Тип",
                ["Держатель душа"],
            )
            if picked:
                inferred.append(("Тип", picked))

    if target_category_id == "2214":
        if "верхн" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Вид лійки",
                ["Верхний душ"],
            )
            if picked:
                inferred.append(("Вид лійки", picked))
        elif "ручн" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Вид лійки",
                ["Ручной душ"],
            )
            if picked:
                inferred.append(("Вид лійки", picked))
        elif "боков" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Вид лійки",
                ["Боковой душ"],
            )
            if picked:
                inferred.append(("Вид лійки", picked))

    if target_category_id == "1429":
        if "диспенсер" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Призначення",
                ["Для антисептика"],
            )
            if picked:
                inferred.append(("Призначення", picked))
        elif "дозатор" in name_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Призначення",
                ["Для мыла"],
            )
            if picked:
                inferred.append(("Призначення", picked))

    return inferred


def infer_forced_attrs_from_offer(
    offer: ET._Element,
    source_category_id: str,
    target_category_id: str,
    merchant_catalog: dict[str, dict],
) -> list[tuple[str, str]]:
    inferred: list[tuple[str, str]] = []

    if target_category_id != "1899":
        return inferred

    mount_value = (
        find_param_value(offer, "Монтаж")
        or find_param_value(offer, "Установка")
        or find_param_value(offer, "Способ монтажа")
        or ""
    )
    mount_key = normalize_key(mount_value)

    if mount_key:
        # Known source installation variants that Maudau expects as "Врезной (на изделие)".
        if any(
            token in mount_key
            for token in (
                "одно отверст",
                "два отверст",
                "три отверст",
                "четыре отверст",
                "на борт ванны",
                "приставка для унитаза",
            )
        ):
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Монтаж",
                ["Врезной (на изделие)"],
            )
            if picked:
                inferred.append(("Монтаж", picked))

        holes_value = ""
        if "одно отверст" in mount_key:
            holes_value = "1 отверстие"
        elif "два отверст" in mount_key:
            holes_value = "2 отверстия"
        elif "три отверст" in mount_key:
            holes_value = "3 отверстия"
        elif "четыре отверст" in mount_key:
            holes_value = "4 отверстия"
        if holes_value:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Кількість отворів",
                [holes_value],
            )
            if picked:
                inferred.append(("Кількість отворів", picked))

        # Rule requested: washbasin mixers + wall mounting -> hidden installation.
        if source_category_id == "1069" and "настенн" in mount_key:
            picked = pick_allowed_value(
                merchant_catalog,
                target_category_id,
                "Встановлення",
                ["Скрытая"],
            )
            if picked:
                inferred.append(("Встановлення", picked))

    # For bidet mixers default spout type is stationary (only if allowed by Maudau).
    if source_category_id == "1071":
        picked = pick_allowed_value(
            merchant_catalog,
            target_category_id,
            "Вилив",
            ["Стационарный"],
        )
        if picked:
            inferred.append(("Вилив", picked))

    return inferred


def collect_forced_attrs_for_source(
    offer: ET._Element,
    source_category_id: str,
    source_category_name: str,
    target_category_id: str,
    merchant_catalog: dict[str, dict],
) -> list[tuple[str, str]]:
    rules = list(FORCED_ATTRS_BY_SOURCE_CATEGORY.get(source_category_id, []))
    rules.extend(infer_forced_attrs_from_source_name(source_category_name, target_category_id, merchant_catalog))
    rules.extend(infer_forced_attrs_from_offer(offer, source_category_id, target_category_id, merchant_catalog))

    towel_type = TOWEL_TYPE_BY_SOURCE_CATEGORY.get(source_category_id)
    if towel_type and target_category_id == "1902":
        rules.append(("Тип", towel_type))

    meta = merchant_catalog.get(target_category_id, {})
    attr_lookup = meta.get("attr_lookup", {})
    attrs = meta.get("attrs", {})
    result: list[tuple[str, str]] = []
    seen: set[tuple[str, str]] = set()

    for raw_name, raw_value in rules:
        mapped_name = map_param_name(raw_name, target_category_id)
        canonical_name = attr_lookup.get(normalize_key(mapped_name), mapped_name)
        normalized_value = apply_category_value_override(target_category_id, canonical_name, raw_value)
        allowed_values = attrs.get(canonical_name, {})
        canonical_value = map_param_value_to_allowed(normalized_value, allowed_values)
        if allowed_values and normalize_text_key(canonical_value) not in allowed_values:
            continue
        key = (normalize_key(canonical_name), normalize_key(canonical_value))
        if not canonical_name or not canonical_value or key in seen:
            continue
        seen.add(key)
        result.append((canonical_name, canonical_value))

    return result


def apply_category_value_override(target_category_id: str, attr_name: str, value: str) -> str:
    clean = compact_text(value)
    value_map = CATEGORY_ATTR_VALUE_OVERRIDES.get(target_category_id, {}).get(attr_name, {})
    if not value_map:
        return clean

    raw_key = normalize_key(clean)
    if raw_key in value_map:
        return value_map[raw_key]

    # Handle combined values like "рычаг|кнопочный" or "рычаг / кнопочный".
    tokens = [t for t in re.split(r"[|/,;+]", raw_key) if t.strip()]
    for token in tokens:
        normalized = value_map.get(token.strip())
        if normalized:
            return normalized

    return clean


def normalize_text_key(value: str) -> str:
    key = normalize_key(value)
    key = key.replace("ё", "е").replace("’", "'").replace("`", "'")
    key = re.sub(r"\s+", " ", key)
    return key.strip()


def normalize_unit_token(unit: str) -> str:
    u = normalize_text_key(unit).replace(".", "")
    replacements = {
        "вт": "вт",
        "квт": "квт",
        "мм": "мм",
        "см": "см",
        "м": "м",
        "мес": "мес",
        "месяц": "мес",
        "месяца": "мес",
        "месяцев": "мес",
        "год": "год",
        "года": "год",
        "годов": "год",
        "м²": "м2",
        "м2": "м2",
    }
    return replacements.get(u, u)


def convert_unit_value(value: float, from_unit: str, to_unit: str) -> float | None:
    f = normalize_unit_token(from_unit)
    t = normalize_unit_token(to_unit)
    if not f and t:
        return value
    if f and not t:
        return value
    if not f or not t or f == t:
        return value
    if f == "мм" and t == "см":
        return value / 10.0
    if f == "см" and t == "мм":
        return value * 10.0
    if f == "вт" and t == "квт":
        return value / 1000.0
    if f == "квт" and t == "вт":
        return value * 1000.0
    if f == "мес" and t == "год":
        return value / 12.0
    if f == "год" and t == "мес":
        return value * 12.0
    return None


def parse_scalar_with_unit(text: str) -> tuple[float, str] | None:
    m = SCALAR_WITH_UNIT_RE.match(compact_text(text))
    if not m:
        return None
    try:
        num = float(m.group(1).replace(",", "."))
    except ValueError:
        return None
    unit = normalize_unit_token(m.group(2) or "")
    return num, unit


def parse_allowed_range(text: str) -> tuple[str, float, float, str] | None:
    raw = compact_text(text)
    m = RANGE_SIMPLE_RE.match(raw)
    if m:
        lo = float(m.group(1).replace(",", "."))
        hi = float(m.group(2).replace(",", "."))
        unit = normalize_unit_token(m.group(3) or "")
        return "between", lo, hi, unit
    m = RANGE_UPPER_RE.match(raw)
    if m:
        hi = float(m.group(2).replace(",", "."))
        unit = normalize_unit_token(m.group(3) or "")
        return "upper", float("-inf"), hi, unit
    m = RANGE_LOWER_RE.match(raw)
    if m:
        lo = float(m.group(2).replace(",", "."))
        unit = normalize_unit_token(m.group(3) or "")
        return "lower", lo, float("inf"), unit
    return None


def fits_allowed_range(source_value: str, allowed_value: str) -> bool:
    scalar = parse_scalar_with_unit(source_value)
    allowed = parse_allowed_range(allowed_value)
    if scalar is None or allowed is None:
        return False
    sval, sunit = scalar
    kind, lo, hi, aunit = allowed
    if aunit:
        converted = convert_unit_value(sval, sunit, aunit)
        if converted is None:
            return False
        sval = converted
    if kind == "between":
        return lo <= sval <= hi
    if kind == "upper":
        return sval <= hi
    if kind == "lower":
        return sval >= lo
    return False


def apply_generic_value_synonyms(value: str) -> str:
    key = normalize_text_key(value)
    if key in GENERIC_VALUE_SYNONYMS:
        return GENERIC_VALUE_SYNONYMS[key]
    return value


def map_param_value_to_allowed(value: str, allowed_values: dict[str, str]) -> str:
    clean = compact_text(value)
    if not clean or not allowed_values:
        return clean

    # Remove finish suffixes globally before matching:
    # matte and glossy words should not change base color semantics.
    clean_base = re.sub(
        r"\b(матов(?:ый|ая|ое|ые|ого|ому|ым|ом|ую|а|о|і|ий)?|глянцев(?:ый|ая|ое|ые|ого|ому|ым|ом|ую)?)\b",
        "",
        clean,
        flags=re.IGNORECASE,
    )
    clean_base = compact_text(clean_base)
    if clean_base:
        clean = clean_base

    direct = allowed_values.get(normalize_text_key(clean))
    if direct:
        return direct

    # Color simplification (e.g., "Черный матовый" -> "Черный") when such base value exists in allowed set.
    color_aliases = {
        "черный": "Черный",
        "черний": "Черный",
        "чорний": "Черный",
        "чорный": "Черный",
        "белый": "Белый",
        "білий": "Белый",
        "серый": "Серый",
        "сірий": "Серый",
        "графит": "Графит",
        "графіт": "Графит",
        "бежевый": "Бежевый",
        "бежевий": "Бежевый",
        "коричневый": "Коричневый",
        "коричневий": "Коричневый",
        "хром": "Хром",
        "хромированный": "Хром",
        "хромована": "Хром",
        "сатин": "Хром",
    }
    key_clean = normalize_text_key(clean).replace("-", " ")
    for token, canonical in color_aliases.items():
        if token in key_clean:
            m = allowed_values.get(normalize_text_key(canonical))
            if m:
                return m

    # 1) Apply generic morphology/language synonyms (RU/UA forms).
    synonym = apply_generic_value_synonyms(clean)
    if synonym != clean:
        direct_syn = allowed_values.get(normalize_text_key(synonym))
        if direct_syn:
            return direct_syn

    for candidate in unit_value_candidates(clean):
        match = allowed_values.get(normalize_text_key(candidate))
        if match:
            return match

    # 2) Exact numeric matching has higher priority than ranges.
    scalar = parse_scalar_with_unit(clean)
    if scalar is not None:
        source_num, source_unit = scalar
        best_exact: tuple[int, str] | None = None
        for allowed_canonical in allowed_values.values():
            if parse_allowed_range(allowed_canonical) is not None:
                continue
            allowed_scalar = parse_scalar_with_unit(allowed_canonical)
            if allowed_scalar is None:
                continue
            target_num, target_unit = allowed_scalar
            converted = convert_unit_value(source_num, source_unit, target_unit)
            if converted is None:
                continue
            if abs(converted - target_num) > 1e-9:
                continue
            score = 2 if normalize_unit_token(source_unit) == normalize_unit_token(target_unit) else 1
            if best_exact is None or score > best_exact[0]:
                best_exact = (score, allowed_canonical)
        if best_exact is not None:
            return best_exact[1]

    # 3) Range-safe mapping: map scalar into allowed ranges only on clear hit.
    for allowed_raw, allowed_canonical in allowed_values.items():
        _ = allowed_raw
        if fits_allowed_range(clean, allowed_canonical):
            return allowed_canonical

    return clean


def infer_color_from_name(name_text: str, allowed_values: dict[str, str]) -> str:
    if not name_text or not allowed_values:
        return ""

    base = compact_text(name_text)
    base = re.sub(
        r"\b(матов(?:ый|ая|ое|ые|ого|ому|ым|ом|ую|а|о|і|ий)?|глянцев(?:ый|ая|ое|ые|ого|ому|ым|ом|ую)?)\b",
        "",
        base,
        flags=re.IGNORECASE,
    )
    name_norm = " " + normalize_text_key(base).replace("-", " ") + " "

    allowed_uniques = list(dict.fromkeys(allowed_values.values()))
    allowed_uniques.sort(key=lambda x: len(normalize_text_key(x)), reverse=True)

    for canonical in allowed_uniques:
        cand = " " + normalize_text_key(canonical).replace("-", " ") + " "
        if cand in name_norm:
            return canonical
    return ""


def strict_map_value_for_attr(
    target_category_id: str,
    attr_name: str,
    value: str,
    merchant_catalog: dict[str, dict],
) -> str:
    meta = merchant_catalog.get(target_category_id, {})
    attrs = meta.get("attrs", {})
    allowed_values = attrs.get(attr_name, {})
    mapped = map_param_value_to_allowed(value, allowed_values)
    if not allowed_values:
        return mapped
    return allowed_values.get(normalize_text_key(mapped), "")


def _params_lookup(source_params: dict[str, str]) -> dict[str, str]:
    return {normalize_key(k): compact_text(v) for k, v in source_params.items() if k and v}


def _find_param(source_params_lut: dict[str, str], *names: str) -> str:
    for n in names:
        v = source_params_lut.get(normalize_key(n), "")
        if v:
            return v
    return ""


def apply_cross_rules(
    target_category_id: str,
    source_category_id: str,
    source_category_name: str,
    source_params: dict[str, str],
    merchant_catalog: dict[str, dict],
) -> dict[str, str]:
    """Apply explicit cross-mapping rules from 'хар. пересечение.xlsx'."""
    out: dict[str, str] = {}
    lut = _params_lookup(source_params)
    meta = merchant_catalog.get(target_category_id, {})
    attrs = meta.get("attrs", {})

    def assign(attr_candidates: list[str], value: str) -> None:
        for attr in attr_candidates:
            if attr not in attrs:
                continue
            mapped = strict_map_value_for_attr(target_category_id, attr, value, merchant_catalog)
            if mapped:
                out[attr] = mapped
                return

    # Section 1: "смесители наши - maudau" (target category 1899).
    if target_category_id == "1899":
        mont = _find_param(lut, "Монтаж", "монтаж", "Способ монтажа", "Тип установки")
        osob = _find_param(lut, "Особенности", "особенности", "Оснащение", "Оснащення")
        nazn = _find_param(lut, "Назначение", "назначение", "Призначення")
        mont_k = normalize_text_key(mont)
        osob_k = normalize_text_key(osob)
        nazn_k = normalize_text_key(nazn)
        is_kitchen_mixer = source_category_id in {"1073", "1224"}

        # Монтаж/установка.
        mount_is_insert = any(
            x in mont_k
            for x in ["на одно отверстие", "на три отверстия", "на борт ванны", "на два отверстия", "приставка для унитаза"]
        )
        if mount_is_insert:
            assign(["Монтаж", "Встановлення", "Установка"], "Врезной (на изделие)")
        elif any(x in mont_k for x in ["наружный", "настенный"]):
            assign(["Монтаж", "Встановлення", "Установка"], "настенный")
        elif "наполь" in mont_k:
            assign(["Монтаж", "Встановлення", "Установка"], "напольный")

        # Внешняя/Скрытая (fallback rules).
        if source_category_id == "1069" and "настенн" in mont_k:
            assign(["Монтаж", "Встановлення"], "Скрытая")
        elif "скрыт" in mont_k:
            assign(["Монтаж", "Встановлення"], "Скрытая")
        elif mont_k and not mount_is_insert:
            assign(["Монтаж", "Встановлення"], "Внешняя")

        # Вилив.
        if any(x in osob_k for x in ["на две воды/гибкий излив", "гибкий излив"]):
            assign(["Вилив"], "Гибкий")
        elif any(x in osob_k for x in ["на две воды/выдвижной излив", "выдвижной излив"]):
            assign(["Вилив"], "Выдвижной")
        elif "каскад" in osob_k:
            assign(["Вилив"], "Каскадный")
        elif "излив" in osob_k:
            assign(["Вилив"], "Поворотный")
        else:
            assign(["Вилив"], "Стационарный")

        # Підключення до фільтру води.
        if is_kitchen_mixer:
            if any(x in osob_k for x in ["на две воды", "на две воды/выдвижной излив", "на две воды/гибкий излив"]):
                assign(["Підключення до фільтру води"], "Есть")
            else:
                assign(["Підключення до фільтру води"], "Нет")

        # Кількість отворів.
        if "на одно отверстие" in mont_k:
            assign(["Кількість отворів"], "1 отверстие")
        elif "на два отверстия" in mont_k:
            assign(["Кількість отворів"], "2 отверстия")
        elif "на три отверстия" in mont_k:
            assign(["Кількість отворів"], "3 отверстия")

        # Оснащення.
        if "с душевой лейкой" in nazn_k or "с душевым гарнитуром" in nazn_k:
            assign(["Оснащення"], "С лейкой для душа")

    # Section 2: "Все для принятия душа" (1904 / 2214 / 2299).
    if target_category_id in {"1904", "2214", "2299"}:
        modes = _find_param(lut, "Количество режимов", "количество режимов")
        dia = _find_param(lut, "Диаметр душа, мм", "диаметр душа, мм")
        hose = _find_param(lut, "Длина душевого шланга, см", "длина душевого шланга, см", "Длина шланга, см")
        mont = _find_param(lut, "Монтаж", "монтаж")

        modes_k = normalize_text_key(modes)
        dia_k = normalize_text_key(dia)
        mont_k = normalize_text_key(mont)

        # Кількість режимів струменя ручного душу.
        m = re.search(r"\d+", modes_k)
        if m:
            assign(["Кількість режимів струменя ручного душу", "Кількість режимів"], m.group(0))

        # Діаметр ручної лійки.
        if "100-199" in dia_k:
            assign(["Діаметр ручної лійки", "Діаметр ручної лійки, мм"], "10 см")
        elif "менее 100" in dia_k or "до 100" in dia_k:
            assign(["Діаметр ручної лійки", "Діаметр ручної лійки, мм"], "8 см")

        # Довжина душового шлангу.
        hose_num = None
        if hose:
            m_h = re.search(r"\d+(?:[.,]\d+)?", normalize_text_key(hose))
            if m_h:
                try:
                    hose_num = float(m_h.group(0).replace(",", "."))
                except ValueError:
                    hose_num = None
        if hose_num is not None:
            # If source field is in centimeters (common case), keep exact cm first.
            if hose_num <= 300:
                assign(["Довжина душового шлангу"], f"{normalize_number(hose_num)} см")
            else:
                # Legacy mapping where source value can be in millimeters.
                if hose_num < 1250:
                    assign(["Довжина душового шлангу"], "100 см")
                elif abs(hose_num - 1250) < 1e-9:
                    assign(["Довжина душового шлангу"], "125 см")
                elif abs(hose_num - 1500) < 1e-9:
                    assign(["Довжина душового шлангу"], "150 см")
                elif abs(hose_num - 1600) < 1e-9:
                    assign(["Довжина душового шлангу"], "160 см")
                elif abs(hose_num - 1700) < 1e-9:
                    assign(["Довжина душового шлангу"], "170 см")
                elif abs(hose_num - 1750) < 1e-9:
                    assign(["Довжина душового шлангу"], "175 см")
                elif abs(hose_num - 1800) < 1e-9:
                    assign(["Довжина душового шлангу"], "180 см")
                elif abs(hose_num - 2000) < 1e-9:
                    assign(["Довжина душового шлангу"], "200 см")

        # Монтаж.
        if "наружн" in mont_k:
            assign(["Монтаж", "Встановлення"], "Настенный")
        elif "скрыт" in mont_k:
            assign(["Монтаж", "Встановлення"], "Скрытый")

    return out


def cleanup_params(offer: ET._Element, target_category_id: str, merchant_catalog: dict[str, dict]) -> None:
    category_meta = merchant_catalog.get(target_category_id, {})
    attr_lookup = category_meta.get("attr_lookup", {})
    attrs = category_meta.get("attrs", {})

    dedupe: set[tuple[str, str]] = set()
    for p in list(offer.findall("param")):
        pname = normalize_text(HTML_TAG_RE.sub("", p.get("name") or ""))
        pval = compact_text(HTML_TAG_RE.sub(" ", p.text or ""))

        if not pname or not pval:
            offer.remove(p)
            continue

        mapped_name = map_param_name(pname, target_category_id)
        canonical_name = attr_lookup.get(normalize_key(mapped_name), mapped_name)
        if attrs and canonical_name not in attrs:
            offer.remove(p)
            continue
        p.set("name", canonical_name)

        allowed_values = attrs.get(canonical_name, {})
        normalized_value = apply_category_value_override(target_category_id, canonical_name, pval)
        mapped_value = map_param_value_to_allowed(normalized_value, allowed_values)
        if allowed_values and normalize_text_key(mapped_value) not in allowed_values:
            offer.remove(p)
            continue
        p.text = mapped_value

        dedupe_key = (normalize_key(canonical_name), normalize_key(p.text))
        if dedupe_key in dedupe:
            offer.remove(p)
            continue
        dedupe.add(dedupe_key)


def apply_forced_category_params(
    offer: ET._Element,
    source_category_id: str,
    source_category_name: str,
    target_category_id: str,
    merchant_catalog: dict[str, dict],
) -> int:
    rules = collect_forced_attrs_for_source(
        offer,
        source_category_id,
        source_category_name,
        target_category_id,
        merchant_catalog,
    )
    if not rules:
        return 0

    changed = 0

    for canonical_name, canonical_value in rules:
        if upsert_param(offer, canonical_name, canonical_value):
            changed += 1

    return changed


def cleanup_pictures(offer: ET._Element) -> None:
    pictures = [p for p in offer.findall("picture")]
    kept = 0
    for pic in pictures:
        url = normalize_text(pic.text)
        if not url or " " in url or len(url) > 255 or CYRILLIC_RE.search(url):
            offer.remove(pic)
            continue
        pic.text = url
        kept += 1
        if kept > 12:
            offer.remove(pic)


def normalize_offer_id(offer: ET._Element) -> bool:
    raw = resolve_offer_id_raw(offer)
    clean = ID_CLEAN_RE.sub("", raw)
    if not clean:
        return False
    offer.set("id", clean)
    return True


def has_required_fields(offer: ET._Element) -> bool:
    required_tags = ["name_ua", "name_ru", "description_ua", "description_ru", "price", "categoryId"]
    for tag in required_tags:
        if not child_text(offer, tag):
            return False

    pics = [p for p in offer.findall("picture") if normalize_text(p.text)]
    return len(pics) > 0


def resolve_target_category_id(offer: ET._Element, source_id: str) -> str:
    target_id = SOURCE_TO_MAUDAU_CATEGORY.get(source_id, source_id)

    # 1167 "Аксессуары": TЭНы -> 3175, остальные -> 3189.
    if source_id == "1167":
        probe = " ".join(
            [
                child_text(offer, "name_ru"),
                child_text(offer, "name_ua"),
                child_text(offer, "description_ru"),
                child_text(offer, "description_ua"),
                find_param_value(offer, "Вид"),
                find_param_value(offer, "Тип"),
            ]
        )
        probe_key = normalize_key(probe)
        if any(token in probe_key for token in ("тэн", "тен", "teh", "ten")):
            return "3175"
        return "3189"

    # 1169 "Комплектующие": only siphons are remapped to 3172.
    if source_id == "1169":
        kind = normalize_key(find_param_value(offer, "Вид"))
        if "сифон" in kind:
            return "3172"
        return source_id

    return target_id


def remap_offer_category(
    offer: ET._Element,
    merchant_catalog: dict[str, dict],
    source_category_names: dict[str, str],
) -> tuple[bool, str, str, int, bool]:
    source_id = child_text(offer, "categoryId")
    if not source_id:
        return False, source_id, "", 0, False

    if source_id in SKIP_REMAP_SOURCE_CATEGORIES or source_id in QUESTION_SOURCE_CATEGORIES:
        # Keep category as-is by explicit business rule.
        return True, source_id, source_id, 0, source_id in merchant_catalog

    target_id = resolve_target_category_id(offer, source_id)
    target_known = target_id in merchant_catalog

    set_or_create(offer, "categoryId", target_id)
    source_name = source_category_names.get(source_id, "")
    forced_changes = 0
    if target_known:
        forced_changes = apply_forced_category_params(offer, source_id, source_name, target_id, merchant_catalog)
    return True, source_id, target_id, forced_changes, target_known


def normalize_offer(
    offer: ET._Element,
    target_category_id: str,
    merchant_catalog: dict[str, dict],
    brands_catalog: dict[str, str],
    countries_catalog: dict[str, str],
) -> bool:
    normalize_name_description(offer)
    normalize_old_price(offer)
    enrich_vendor_country_from_params(offer)
    normalize_vendor_by_catalog(offer, brands_catalog)
    normalize_country_by_catalog(offer, countries_catalog)
    cleanup_params(offer, target_category_id, merchant_catalog)
    cleanup_pictures(offer)
    offer.attrib.pop("group_id", None)

    url_node = offer.find("url")
    if url_node is not None:
        offer.remove(url_node)

    if not normalize_offer_id(offer):
        return False

    set_available(offer, extract_available(offer))

    return has_required_fields(offer)


def ensure_root_date(root: ET._Element) -> None:
    # Always refresh generation timestamp to match current feed snapshot.
    root.set("date", datetime.now().strftime("%Y-%m-%d %H:%M"))


def ensure_unique_offer_ids(root: ET._Element) -> int:
    """Ensure all offer ids are unique and remain [A-Za-z0-9]."""
    used: set[str] = set()
    changed = 0
    for offer in root.xpath("//offer"):
        base = normalize_text(offer.get("id"))
        if not base:
            continue
        if base not in used:
            used.add(base)
            continue

        suffix = 2
        candidate = f"{base}{suffix}"
        while candidate in used:
            suffix += 1
            candidate = f"{base}{suffix}"
        offer.set("id", candidate)
        used.add(candidate)
        changed += 1
    return changed


def rebuild_categories(
    root: ET._Element,
    merchant_catalog: dict[str, dict],
    source_category_names: dict[str, str],
) -> None:
    shop = root.find("shop")
    if shop is None:
        shop = ET.SubElement(root, "shop")

    existing = shop.find("categories")
    if existing is not None:
        shop.remove(existing)

    categories = ET.Element("categories")
    known = set()

    def add_category(cid: str) -> None:
        if not cid or cid in known:
            return
        c = ET.SubElement(categories, "category", id=cid)
        c.text = (
            MAUDAU_CATEGORY_NAME_OVERRIDES.get(cid)
            or merchant_catalog.get(cid, {}).get("name_ru")
            or source_category_names.get(cid)
            or cid
        )
        known.add(cid)

    for offer in shop.xpath(".//offers/offer"):
        cid = child_text(offer, "categoryId")
        add_category(cid)

    # Keep mapped categories visible in header even when current offer slice is empty after filtering.
    for sid in source_category_names.keys():
        if sid in SKIP_REMAP_SOURCE_CATEGORIES or sid in QUESTION_SOURCE_CATEGORIES:
            continue

        extra_ids: list[str] = []
        if sid == "1167":
            extra_ids = ["3175", "3189"]
        elif sid == "1169":
            extra_ids = ["3172"]
        else:
            tid = SOURCE_TO_MAUDAU_CATEGORY.get(sid, sid)
            if tid and tid != sid:
                extra_ids = [tid]

        for tid in extra_ids:
            if tid in merchant_catalog or tid in MAUDAU_CATEGORY_NAME_OVERRIDES:
                add_category(tid)

    shop.insert(0, categories)


def build_rozetka_index(tree: ET._ElementTree) -> dict[str, dict[str, str]]:
    index = {}
    for offer in tree.xpath("//offer"):
        key = resolve_offer_id_key(offer)
        if not key or key in index:
            continue
        index[key] = {
            "price": child_text(offer, "price"),
            "old_price": extract_old_price(offer),
            "available": extract_available(offer),
        }
    return index


def build_gap_report(source_root: ET._Element, merchant_catalog: dict[str, dict], report_path: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        print("⚠ openpyxl не установлен, отчет по непересечениям пропущен")
        return

    source_category_name = {
        normalize_text(c.get("id")): normalize_text(c.text)
        for c in source_root.xpath("//shop/categories/category")
    }

    offers_by_source: Counter[str] = Counter()
    source_param_counts: dict[str, Counter[str]] = defaultdict(Counter)
    source_param_values: dict[str, dict[str, Counter[str]]] = defaultdict(lambda: defaultdict(Counter))

    for offer in source_root.xpath("//offer"):
        source_id = child_text(offer, "categoryId")
        if not source_id:
            continue
        offers_by_source[source_id] += 1
        for param in offer.findall("param"):
            name = normalize_text(param.get("name"))
            value = compact_text(param.text or "")
            if not name or not value:
                continue
            source_param_counts[source_id][name] += 1
            source_param_values[source_id][name][value] += 1

    mapped_rows = []
    unmapped_categories_rows = []
    param_name_gap_rows = []
    param_value_gap_rows = []
    maudau_missing_attr_rows = []

    used_source_categories = sorted(k for k, v in offers_by_source.items() if v > 0)

    for source_id in used_source_categories:
        source_name = source_category_name.get(source_id, "")
        offers_count = offers_by_source[source_id]
        target_id = SOURCE_TO_MAUDAU_CATEGORY.get(source_id, source_id)
        target_meta = merchant_catalog.get(target_id)

        if not target_meta:
            if source_id in SKIP_REMAP_SOURCE_CATEGORIES:
                comment = "Будет создана на Maudau позже (без правок сейчас)"
            elif source_id in QUESTION_SOURCE_CATEGORIES:
                comment = "?"
            else:
                comment = "Нет категории в merchant_categories.xml"
            unmapped_categories_rows.append([
                source_id,
                source_name,
                offers_count,
                "",
                "",
                comment,
            ])
            continue

        mapped_rows.append([
            source_id,
            source_name,
            offers_count,
            target_id,
            target_meta.get("name_ru", ""),
        ])

        attr_lookup = target_meta.get("attr_lookup", {})
        attrs = target_meta.get("attrs", {})

        covered_attrs: set[str] = set()
        forced_attrs = collect_forced_attrs_for_source(
            source_id,
            source_name,
            target_id,
            merchant_catalog,
        )
        for forced_attr_name, _forced_value in forced_attrs:
            covered_attrs.add(forced_attr_name)

        for source_param_name, cnt in sorted(source_param_counts[source_id].items(), key=lambda x: (-x[1], x[0])):
            mapped_name = map_param_name(source_param_name, target_id)
            canonical_name = attr_lookup.get(normalize_key(mapped_name), mapped_name)
            is_attr = canonical_name in attrs

            if is_attr:
                covered_attrs.add(canonical_name)
            else:
                param_name_gap_rows.append([
                    source_id,
                    source_name,
                    target_id,
                    target_meta.get("name_ru", ""),
                    source_param_name,
                    cnt,
                    mapped_name,
                    "Нет атрибута в Maudau категории",
                ])
                continue

            allowed_values = attrs.get(canonical_name, {})
            if not allowed_values:
                continue

            for src_value, value_cnt in source_param_values[source_id][source_param_name].items():
                normalized_direct = normalize_key(src_value)
                if normalized_direct in allowed_values:
                    continue

                matched = False
                for candidate in unit_value_candidates(src_value):
                    if normalize_key(candidate) in allowed_values:
                        matched = True
                        break

                if not matched:
                    allowed_sample = ", ".join(list(dict.fromkeys(allowed_values.values()))[:5])
                    param_value_gap_rows.append([
                        source_id,
                        source_name,
                        target_id,
                        target_meta.get("name_ru", ""),
                        canonical_name,
                        src_value,
                        value_cnt,
                        allowed_sample,
                    ])

        for attr_name in sorted(attrs.keys()):
            if attr_name not in covered_attrs:
                maudau_missing_attr_rows.append([
                    source_id,
                    source_name,
                    target_id,
                    target_meta.get("name_ru", ""),
                    attr_name,
                    "Нет соответствующего source-параметра",
                ])

    wb = Workbook()
    ws_map = wb.active
    ws_map.title = "mapped_categories"
    ws_unmapped = wb.create_sheet("unmapped_categories")
    ws_param_name = wb.create_sheet("param_name_gaps")
    ws_param_value = wb.create_sheet("param_value_gaps")
    ws_attr_miss = wb.create_sheet("maudau_attr_missing")

    ws_map.append(["source_id", "source_category", "offers_count", "maudau_id", "maudau_category"])
    ws_unmapped.append(["source_id", "source_category", "offers_count", "maudau_id", "maudau_category", "comment"])
    ws_param_name.append([
        "source_id",
        "source_category",
        "maudau_id",
        "maudau_category",
        "source_param",
        "offers_with_param",
        "mapped_name_attempt",
        "comment",
    ])
    ws_param_value.append([
        "source_id",
        "source_category",
        "maudau_id",
        "maudau_category",
        "maudau_attr",
        "source_value",
        "offers_with_value",
        "allowed_values_sample",
    ])
    ws_attr_miss.append([
        "source_id",
        "source_category",
        "maudau_id",
        "maudau_category",
        "maudau_attr",
        "comment",
    ])

    for row in mapped_rows:
        ws_map.append(row)
    for row in unmapped_categories_rows:
        ws_unmapped.append(row)
    for row in param_name_gap_rows:
        ws_param_name.append(row)
    for row in param_value_gap_rows:
        ws_param_value.append(row)
    for row in maudau_missing_attr_rows:
        ws_attr_miss.append(row)

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for sheet in [ws_map, ws_unmapped, ws_param_name, ws_param_value, ws_attr_miss]:
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        sheet.freeze_panes = "A2"

    for sheet in [ws_map, ws_unmapped, ws_param_name, ws_param_value, ws_attr_miss]:
        for col in sheet.columns:
            max_len = 0
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            sheet.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 12), 70)

    wb.save(str(report_path))
    print(f"📊 Отчет по непересечениям сохранен: {report_path}")


def parse_template_target(value: object) -> tuple[str, str]:
    if value is None:
        return "empty", ""
    if isinstance(value, (int, float)):
        return "mapped", str(int(value))
    text = normalize_text(str(value))
    if not text:
        return "empty", ""
    if text.isdigit():
        return "mapped", text
    low = normalize_key(text)
    if "будет создана" in low:
        return "later", text
    if text == "?":
        return "question", text
    return "text", text


def build_final_template_report(
    source_root: ET._Element,
    merchant_catalog: dict[str, dict],
    template_path: Path,
    output_path: Path,
) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import PatternFill
    except Exception:
        print("⚠ openpyxl не установлен, финальный шаблонный отчет пропущен")
        return

    if not template_path.exists():
        print(f"⚠ Шаблон финального отчета не найден: {template_path}")
        return

    wb = load_workbook(str(template_path))
    ws = wb["Лист1"] if "Лист1" in wb.sheetnames else wb.active

    source_category_name = {
        normalize_text(c.get("id")): normalize_text(c.text)
        for c in source_root.xpath("//shop/categories/category")
    }
    offers_by_source: Counter[str] = Counter()
    source_param_counts: dict[str, Counter[str]] = defaultdict(Counter)
    source_param_values: dict[str, dict[str, Counter[str]]] = defaultdict(lambda: defaultdict(Counter))

    for offer in source_root.xpath("//offer"):
        source_id = child_text(offer, "categoryId")
        if not source_id:
            continue
        offers_by_source[source_id] += 1
        for param in offer.findall("param"):
            name = normalize_text(param.get("name"))
            value = compact_text(param.text or "")
            if not name or not value:
                continue
            source_param_counts[source_id][name] += 1
            source_param_values[source_id][name][value] += 1

    yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")

    rows: list[dict] = []
    for r in range(3, ws.max_row + 1):
        sid = ws.cell(r, 1).value
        if sid is None or normalize_text(str(sid)) == "":
            continue
        source_id = str(int(sid)) if isinstance(sid, (int, float)) else normalize_text(str(sid))
        target_raw = ws.cell(r, 4).value
        status, target = parse_template_target(target_raw)
        row_values = [ws.cell(r, c).value for c in range(1, 18)]
        rows.append(
            {
                "source_id": source_id,
                "status": status,
                "target": target,
                "raw_target": target_raw,
                "values": row_values,
            }
        )

    def row_key(item: dict) -> tuple[int, str]:
        order = {"mapped": 0, "text": 1, "empty": 2, "later": 3, "question": 4}
        return order.get(item["status"], 5), item["source_id"]

    rows.sort(key=row_key)

    for r in range(3, ws.max_row + 1):
        for c in range(1, 18):
            ws.cell(r, c).value = None
            ws.cell(r, c).comment = None
            ws.cell(r, c).fill = PatternFill(fill_type=None)

    out_row = 3
    for item in rows:
        source_id = item["source_id"]
        target_id = item["target"]
        status = item["status"]

        for c in range(1, 18):
            ws.cell(out_row, c).value = item["values"][c - 1]

        ws.cell(out_row, 1).value = source_id
        ws.cell(out_row, 2).value = source_category_name.get(source_id, ws.cell(out_row, 2).value)
        ws.cell(out_row, 3).value = offers_by_source.get(source_id, 0)

        if status == "mapped":
            if target_id in merchant_catalog:
                target_meta = merchant_catalog[target_id]
                ws.cell(out_row, 5).value = target_meta.get("name_ru", "")
                attr_lookup = target_meta.get("attr_lookup", {})
                attrs = target_meta.get("attrs", {})
                forced = collect_forced_attrs_for_source(
                    source_id,
                    normalize_text(ws.cell(out_row, 2).value),
                    target_id,
                    merchant_catalog,
                )
                forced_attr_names = {name for name, _ in forced}
                src_params = [p for p, _ in source_param_counts[source_id].most_common()]
                matched_src: list[str] = []
                matched_attrs: list[str] = []
                missing_attrs: list[str] = []

                for p in src_params:
                    mapped_name = map_param_name(p, target_id)
                    canonical = attr_lookup.get(normalize_key(mapped_name), mapped_name)
                    if canonical in attrs:
                        matched_src.append(p)
                        if canonical not in matched_attrs:
                            matched_attrs.append(canonical)

                for forced_name in forced_attr_names:
                    if forced_name not in matched_attrs:
                        matched_attrs.append(forced_name)

                for attr_name in attrs.keys():
                    if attr_name not in matched_attrs:
                        missing_attrs.append(attr_name)

                ws.cell(out_row, 6).value = ", ".join(matched_src[:8])
                ws.cell(out_row, 11).value = ", ".join(matched_attrs[:8])
                ws.cell(out_row, 12).value = ", ".join(missing_attrs[:8])
                ws.cell(out_row, 7).value = ", ".join([f"{k}={v}" for k, v in forced[:4]])

                if matched_src:
                    ws.cell(out_row, 6).fill = yellow
                if matched_attrs:
                    ws.cell(out_row, 11).fill = yellow

                if missing_attrs:
                    lines = []
                    for attr_name in missing_attrs[:6]:
                        vals = list(dict.fromkeys(attrs.get(attr_name, {}).values()))
                        preview = ", ".join(vals[:10])
                        lines.append(f"{attr_name}: {preview}")
                    ws.cell(out_row, 12).comment = Comment("\n".join(lines), "Codex")
            else:
                ws.cell(out_row, 5).value = "ID не найден в merchant_categories"
        elif status == "later":
            ws.cell(out_row, 4).value = "будет создана на maudau позже (пока не делаем с ней правок)"
        elif status == "question":
            ws.cell(out_row, 4).value = "?"

        out_row += 1

    wb.save(str(output_path))
    print(f"📋 Финальный унифицированный отчет сохранен: {output_path}")


def build_offer_detail_sheet(
    source_root: ET._Element,
    merchant_catalog: dict[str, dict],
    output_path: Path,
) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except Exception:
        print("⚠ openpyxl не установлен, лист 'Товары' не создан")
        return

    if not output_path.exists():
        print(f"⚠ Файл не найден для добавления листа товаров: {output_path}")
        return

    wb = load_workbook(str(output_path))
    if "Товары" in wb.sheetnames:
        del wb["Товары"]
    ws = wb.create_sheet("Товары")

    headers = [
        "Артикул (offer)",
        "Название RU",
        "Описание RU",
        "source_category_id",
        "source_category",
        "maudau_category_id",
        "maudau_category",
        "status",
    ]
    ws.append(headers)

    source_category_names = build_source_category_names(source_root)

    for offer in source_root.xpath("//offer"):
        source_id = child_text(offer, "categoryId")
        target_id = SOURCE_TO_MAUDAU_CATEGORY.get(source_id, source_id)
        source_name = source_category_names.get(source_id, "")

        if source_id in SKIP_REMAP_SOURCE_CATEGORIES:
            status = "будет создана на maudau позже (пока не делаем с ней правок)"
        elif source_id in QUESTION_SOURCE_CATEGORIES:
            status = "?"
        elif target_id in merchant_catalog:
            status = "mapped"
        else:
            status = "unmapped"

        article = resolve_offer_id_raw(offer)
        name_ru = child_text(offer, "name_ru") or child_text(offer, "name")
        description_ru = child_text(offer, "description_ru") or child_text(offer, "description")
        maudau_name = merchant_catalog.get(target_id, {}).get("name_ru", "")

        ws.append(
            [
                article,
                name_ru,
                description_ru,
                source_id,
                source_name,
                target_id,
                maudau_name,
                status,
            ]
        )

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    widths = {"A": 22, "B": 44, "C": 80, "D": 16, "E": 34, "F": 16, "G": 32, "H": 18}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(str(output_path))
    print(f"📦 Добавлен лист с товарами: {output_path}#Товары")


def build_single_sheet_offer_report(
    source_root: ET._Element,
    merchant_catalog: dict[str, dict],
    external_category_names: dict[str, str],
    template_path: Path,
    output_path: Path,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.comments import Comment
        from openpyxl.worksheet.datavalidation import DataValidation
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        print("⚠ openpyxl не установлен, единый отчет не создан")
        return

    def strict_allowed_value(value: str, allowed_values: dict[str, str]) -> str:
        clean = compact_text(value)
        if not clean:
            return ""
        if not allowed_values:
            return clean
        mapped = map_param_value_to_allowed(clean, allowed_values)
        if not mapped:
            return ""
        return allowed_values.get(normalize_text_key(mapped), "")

    source_category_names = build_source_category_names(source_root)
    source_category_order = [normalize_text(c.get("id")) for c in source_root.xpath("//shop/categories/category") if normalize_text(c.get("id"))]
    order_index = {sid: idx for idx, sid in enumerate(source_category_order)}

    offers_rows: list[dict] = []
    source_param_counts: dict[str, Counter[str]] = defaultdict(Counter)
    source_param_values: dict[str, dict[str, str]] = defaultdict(dict)
    source_offers_present: set[str] = set()

    for offer in source_root.xpath("//offer"):
        source_id = child_text(offer, "categoryId")
        source_name = source_category_names.get(source_id, "")
        source_offers_present.add(source_id)

        params: dict[str, str] = {}
        for p in offer.findall("param"):
            pname = normalize_text(p.get("name"))
            pval = compact_text(p.text or "")
            if not pname or not pval:
                continue
            if pname not in params:
                params[pname] = pval
            source_param_counts[source_id][pname] += 1
            source_param_values[source_id].setdefault(normalize_text_key(pname), pname)

        offers_rows.append(
            {
                "article": resolve_offer_id_raw(offer),
                "name_ru": child_text(offer, "name_ru") or child_text(offer, "name"),
                "desc_ru": child_text(offer, "description_ru") or child_text(offer, "description"),
                "source_id": source_id,
                "source_name": source_name,
                "params": params,
            }
        )

    # Build category groups: several source categories -> one Maudau category block.
    mapped_groups: dict[str, list[str]] = defaultdict(list)
    later_sources: list[str] = []
    question_sources: list[str] = []

    for sid in source_category_order:
        if sid not in source_offers_present and sid not in FORCE_LAYOUT_SOURCE_CATEGORIES:
            continue
        if sid in SKIP_REMAP_SOURCE_CATEGORIES:
            later_sources.append(sid)
            continue
        if sid in QUESTION_SOURCE_CATEGORIES:
            question_sources.append(sid)
            continue
        if sid not in SOURCE_TO_MAUDAU_CATEGORY:
            question_sources.append(sid)
            continue
        tid = SOURCE_TO_MAUDAU_CATEGORY[sid]
        mapped_groups[tid].append(sid)

    def sort_sids(items: list[str]) -> list[str]:
        return sorted(items, key=lambda s: order_index.get(s, 10**9))

    for k in list(mapped_groups.keys()):
        mapped_groups[k] = sort_sids(mapped_groups[k])
    later_sources = sort_sids(later_sources)
    question_sources = sort_sids(question_sources)

    # Source params by category.
    source_headers: dict[str, list[str]] = {}
    for sid in set(source_offers_present).union(FORCE_LAYOUT_SOURCE_CATEGORIES):
        headers = [p for p, _ in source_param_counts[sid].most_common()]
        if not headers:
            headers = ["(нет параметров)"]
        source_headers[sid] = headers

    # Maudau attrs by category.
    maudau_headers: dict[str, list[str]] = {}
    for tid in mapped_groups:
        meta = merchant_catalog.get(tid, {})
        maudau_headers[tid] = list(meta.get("attrs", {}).keys())

    # Create workbook from scratch (single sheet).
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    yellow = PatternFill(start_color="FFF59D", end_color="FFF59D", fill_type="solid")
    red = PatternFill(start_color="F4B6C2", end_color="F4B6C2", fill_type="solid")
    border = Border(
        left=Side(style="medium", color="000000"),
        right=Side(style="medium", color="000000"),
        top=Side(style="medium", color="000000"),
        bottom=Side(style="medium", color="000000"),
    )

    # Base columns.
    base_headers = ["Артикул", "Название RU", "Описание RU", "category id", "раздел", "category id", "раздел"]
    for i, h in enumerate(base_headers, start=1):
        ws.cell(2, i).value = h

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    ws.cell(1, 1).value = "ИСХОДНИК"
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=7)
    ws.cell(1, 6).value = "MAUDAU"

    # Block layout.
    blocks: list[dict] = []
    col = 8

    mapped_target_order = sorted(
        mapped_groups.keys(),
        key=lambda tid: min(order_index.get(sid, 10**9) for sid in mapped_groups[tid]),
    )

    for tid in mapped_target_order:
        for sid in mapped_groups[tid]:
            headers = source_headers.get(sid, [])
            start = col
            for h in headers:
                ws.cell(2, col).value = h
                col += 1
            end = col - 1
            source_name = source_category_names.get(sid, sid)
            blocks.append(
                {
                    "type": "source",
                    "source_id": sid,
                    "start": start,
                    "end": end,
                    "title": f"{source_name} (наши)",
                    "headers": headers,
                }
            )

        mheaders = maudau_headers.get(tid, [])
        if not mheaders:
            mheaders = ["(параметры не найдены)"]
        start = col
        for h in mheaders:
            ws.cell(2, col).value = h
            col += 1
        end = col - 1
        meta = merchant_catalog.get(tid, {})
        mname = meta.get("name_ru") or external_category_names.get(tid) or f"ID {tid}"
        blocks.append(
            {
                "type": "maudau",
                "target_id": tid,
                "start": start,
                "end": end,
                "title": f"{mname} Maudau",
                "headers": mheaders,
            }
        )

    # Unmapped categories in the end.
    for sid in later_sources + question_sources:
        headers = source_headers.get(sid, [])
        start = col
        for h in headers:
            ws.cell(2, col).value = h
            col += 1
        end = col - 1
        source_name = source_category_names.get(sid, sid)
        blocks.append(
            {
                "type": "source_unmapped",
                "source_id": sid,
                "start": start,
                "end": end,
                "title": f"{source_name} (наши)",
                "headers": headers,
            }
        )

    # Row1 titles + merged cells.
    for block in blocks:
        ws.merge_cells(start_row=1, start_column=block["start"], end_row=1, end_column=block["end"])
        ws.cell(1, block["start"]).value = block["title"]

    max_col = col - 1

    # Header styling.
    for r in [1, 2]:
        for c in range(1, max_col + 1):
            cell = ws.cell(r, c)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Header dropdown lists (allowed Maudau values).
    list_ws = wb.create_sheet("_lists")
    list_ws.sheet_state = "hidden"
    list_col = 1
    maudau_validations: dict[int, object] = {}
    for block in blocks:
        if block["type"] != "maudau":
            continue
        tid = block["target_id"]
        meta = merchant_catalog.get(tid, {})
        attrs = meta.get("attrs", {})
        for offset, header in enumerate(block["headers"]):
            c = block["start"] + offset
            allowed = list(dict.fromkeys(attrs.get(header, {}).values()))
            if not allowed:
                continue
            preview = ", ".join(allowed[:12])
            ws.cell(2, c).comment = Comment(
                f"Категория Maudau: {tid}\nПараметр: {header}\nЗначения (пример): {preview}",
                "Codex",
            )
            col_letter = get_column_letter(list_col)
            for idx, val in enumerate(allowed, start=1):
                list_ws.cell(idx, list_col).value = val
            rng = f"'{list_ws.title}'!${col_letter}$1:${col_letter}${len(allowed)}"
            dv = DataValidation(type="list", formula1=rng, allow_blank=True)
            dv.showDropDown = False
            ws.add_data_validation(dv)
            maudau_validations[c] = dv
            list_col += 1

    # Quick lookup for blocks.
    source_blocks_by_sid: dict[str, dict] = {}
    maudau_block_by_tid: dict[str, dict] = {}
    for block in blocks:
        if block["type"] == "source":
            source_blocks_by_sid[block["source_id"]] = block
        elif block["type"] == "maudau":
            maudau_block_by_tid[block["target_id"]] = block
        elif block["type"] == "source_unmapped":
            source_blocks_by_sid[block["source_id"]] = block

    # Offers ordering.
    offers_rows.sort(key=lambda x: (order_index.get(x["source_id"], 10**9), normalize_key(x["article"])))

    out_row = 3
    for offer in offers_rows:
        sid = offer["source_id"]
        sname = offer["source_name"]
        params = offer["params"]

        if sid in SKIP_REMAP_SOURCE_CATEGORIES:
            target_cell = "будет создана на maudau позже (пока не делаем с ней правок)"
            target_name = ""
            tid = ""
        elif sid in QUESTION_SOURCE_CATEGORIES:
            target_cell = "?"
            target_name = ""
            tid = ""
        else:
            if sid not in SOURCE_TO_MAUDAU_CATEGORY:
                tid = ""
                target_cell = "?"
                target_name = ""
            else:
                tid = SOURCE_TO_MAUDAU_CATEGORY[sid]
                target_cell = tid
                target_name = merchant_catalog.get(tid, {}).get("name_ru", "") or external_category_names.get(tid, "")

        ws.cell(out_row, 1).value = offer["article"]
        ws.cell(out_row, 2).value = offer["name_ru"]
        ws.cell(out_row, 3).value = offer["desc_ru"]
        ws.cell(out_row, 4).value = sid
        ws.cell(out_row, 5).value = sname
        ws.cell(out_row, 6).value = target_cell
        ws.cell(out_row, 7).value = target_name

        # Fill source block.
        sblock = source_blocks_by_sid.get(sid)
        if sblock:
            for i, h in enumerate(sblock["headers"]):
                v = params.get(h, "")
                if v:
                    ws.cell(out_row, sblock["start"] + i).value = v

        # Fill Maudau block (strict allowed mapping + coloring).
        if tid and tid in maudau_block_by_tid and tid in merchant_catalog:
            mblock = maudau_block_by_tid[tid]
            meta = merchant_catalog.get(tid, {})
            attr_lookup = meta.get("attr_lookup", {})
            attrs = meta.get("attrs", {})

            mapped_attrs: dict[str, str] = {}

            for pname, pval in params.items():
                mapped_name = map_param_name(pname, tid)
                canonical_name = attr_lookup.get(normalize_key(mapped_name), mapped_name)
                if canonical_name not in attrs:
                    continue
                normalized = apply_category_value_override(tid, canonical_name, pval)
                strict = strict_allowed_value(normalized, attrs.get(canonical_name, {}))
                if strict and canonical_name not in mapped_attrs:
                    mapped_attrs[canonical_name] = strict

            forced = collect_forced_attrs_for_source(sid, sname, tid, merchant_catalog)
            for k, v in forced:
                allowed_values = attrs.get(k, {})
                strict = strict_allowed_value(v, allowed_values)
                if strict:
                    mapped_attrs[k] = strict

            # Explicit cross-rules from "хар. пересечение.xlsx" logic.
            cross = apply_cross_rules(tid, sid, sname, params, merchant_catalog)
            for k, v in cross.items():
                mapped_attrs[k] = v

            # Fallback color from product name when source color param is non-specific
            # (e.g. "Цветной"), but title contains a concrete allowed color.
            if "Колір" in attrs and not mapped_attrs.get("Колір"):
                inferred_color = infer_color_from_name(offer["name_ru"], attrs.get("Колір", {}))
                strict_color = strict_allowed_value(inferred_color, attrs.get("Колір", {}))
                if strict_color:
                    mapped_attrs["Колір"] = strict_color

            for i, h in enumerate(mblock["headers"]):
                cell = ws.cell(out_row, mblock["start"] + i)
                value = mapped_attrs.get(h, "")
                if value:
                    cell.value = value
                    cell.fill = yellow
                else:
                    cell.value = ""
                    cell.fill = red

        out_row += 1

    # Apply dropdowns to all populated data rows.
    for c in range(1, max_col + 1):
        dv = maudau_validations.get(c)
        if dv is None:
            continue
        col_letter = get_column_letter(c)
        dv.add(f"{col_letter}3:{col_letter}{max(3, out_row - 1)}")

    # Column widths.
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 44
    ws.column_dimensions["C"].width = 64
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 28
    for c in range(8, max_col + 1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = 16

    ws.freeze_panes = "A3"
    wb.save(str(output_path))
    print(f"📋 Единый товарный отчет сохранен: {output_path}")


def main() -> int:
    try:
        print("===== СТАРТ =====")
        sources_state = load_sources_state()
        stale_alerts: list[str] = []

        merchant_categories_paths = resolve_merchant_categories_paths()
        merchant_catalog: dict[str, dict] = {}
        for p in merchant_categories_paths:
            merchant_catalog = merge_merchant_catalogs(merchant_catalog, load_merchant_catalog(p))
        category_list_path = resolve_category_list_path()
        external_category_names = load_category_names_from_xlsx(category_list_path)
        brand_list_path = resolve_brand_list_path()
        brands_catalog = load_brands_catalog(brand_list_path)
        countries_list_path = resolve_countries_list_path()
        countries_catalog = load_countries_catalog(countries_list_path)
        if merchant_categories_paths:
            print("✅ Подключены справочники категорий Maudau:")
            for p in merchant_categories_paths:
                print(f"   - {p}")
        else:
            print("⚠ merchant_categories.xml не найден, строгая проверка категорий и отчет отключены")
        if category_list_path:
            print(f"✅ Подключен список категорий Maudau (XLSX): {category_list_path}")
        if brand_list_path:
            print(f"✅ Подключен список брендов Maudau (XLSX): {brand_list_path} [{len(brands_catalog)}]")
        if countries_list_path:
            print(f"✅ Подключен список стран Maudau (XLSX): {countries_list_path} [{len(countries_catalog)}]")

        rozetka_path = ROZETKA_XML
        rozetka_mode = "download"
        try:
            download_file(ROZETKA_FEED_URL, ROZETKA_XML, "Розетка XML")
            shutil.copy2(ROZETKA_XML, ROZETKA_BACKUP_XML)
            update_source_success(sources_state, "parserbiz", ROZETKA_XML)
        except Exception as rozetka_exc:
            backup = resolve_rozetka_backup_path()
            if backup is None:
                raise rozetka_exc
            rozetka_path = backup
            rozetka_mode = f"local_fallback ({backup})"
            if backup.resolve() != ROZETKA_BACKUP_XML.resolve():
                shutil.copy2(backup, ROZETKA_BACKUP_XML)
            update_source_failure(sources_state, "parserbiz")
            print(f"⚠ Розетка недоступна, используем локальный файл: {backup}")
            alert = stale_alert_text(sources_state, "parserbiz", "Исходник Parser.biz", backup)
            if alert:
                stale_alerts.append(alert)

        base_path = BASE_XML
        base_mode = "download"
        try:
            download_file(BASE_FEED_URL, BASE_XML, "Maudau XML")
            shutil.copy2(BASE_XML, BASE_BACKUP_XML)
            update_source_success(sources_state, "aquafavorit", BASE_XML)
        except Exception as base_exc:
            backup = resolve_base_backup_path()
            if backup is None:
                raise base_exc
            base_path = backup
            base_mode = f"local_fallback ({backup})"
            if backup.resolve() != BASE_BACKUP_XML.resolve():
                shutil.copy2(backup, BASE_BACKUP_XML)
            update_source_failure(sources_state, "aquafavorit")
            print(f"⚠ AquaFavorit недоступен, используем локальный файл: {backup}")
            alert = stale_alert_text(sources_state, "aquafavorit", "Исходник Aquafavorit", backup)
            if alert:
                stale_alerts.append(alert)

        save_sources_state(sources_state)

        rozetka_tree = ET.parse(str(rozetka_path))
        rozetka_idx = build_rozetka_index(rozetka_tree)

        tree = ET.parse(str(base_path))
        root = tree.getroot()
        source_category_names = build_source_category_names(root)

        # XLSX summary/report generation is intentionally disabled in repo version.
        # Feed update logic remains fully enabled below.

        kept = 0
        removed_missing = 0
        removed_invalid = 0
        unresolved_target_category = 0
        removed_unknown_target_category = 0
        changed_price = 0
        changed_other = 0
        changed_category = 0
        changed_params = 0

        offers = root.xpath("//offer")
        for offer in list(offers):
            source_category_id = child_text(offer, "categoryId")
            vendor = normalize_key(child_text(offer, "vendor"))
            key = resolve_offer_id_key(offer)
            rz = rozetka_idx.get(key)

            keep_without_rozetka = source_category_id in KEEP_WITHOUT_ROZETKA_SOURCE_CATEGORIES
            if rz is None and vendor not in ALLOWED_VENDORS and not keep_without_rozetka:
                offer.getparent().remove(offer)
                removed_missing += 1
                continue

            if rz:
                if set_or_create(offer, "price", rz.get("price", "")):
                    changed_price += 1
                if set_or_create(offer, "old_price", rz.get("old_price", "")):
                    changed_other += 1
                if set_available(offer, rz.get("available", "")):
                    changed_other += 1
            else:
                set_available(offer, extract_available(offer))

            remap_ok, _source_id, target_id, forced_changes, target_known = remap_offer_category(
                offer,
                merchant_catalog,
                source_category_names,
            )
            if not remap_ok:
                offer.getparent().remove(offer)
                removed_invalid += 1
                continue
            if merchant_catalog and not target_known:
                # Temporary safe mode: do not export offers for categories
                # that are not yet present in MAUDAU merchant categories.
                offer.getparent().remove(offer)
                unresolved_target_category += 1
                removed_unknown_target_category += 1
                continue

            if source_category_id != child_text(offer, "categoryId"):
                changed_category += 1
            changed_params += forced_changes

            target_category_id = child_text(offer, "categoryId")
            if not normalize_offer(
                offer,
                target_category_id,
                merchant_catalog,
                brands_catalog,
                countries_catalog,
            ):
                offer.getparent().remove(offer)
                removed_invalid += 1
                continue

            kept += 1

        ensure_root_date(root)
        deduped_ids = ensure_unique_offer_ids(root)
        rebuild_categories(root, merchant_catalog, source_category_names)

        tree.write(str(OUTPUT_XML), encoding="UTF-8", xml_declaration=True, pretty_print=False)
        try:
            shutil.copy2(OUTPUT_XML, LOCAL_OUTPUT_XML)
        except Exception as exc:
            print(f"⚠ Не удалось обновить локальную копию XML ({LOCAL_OUTPUT_XML}): {exc}")

        size_mb = OUTPUT_XML.stat().st_size / (1024 * 1024)

        report = f"""===== СТАРТ =====
▶ Загрузка: Розетка XML
✅ Розетка XML загружен
▶ Загрузка: Maudau XML
✅ Maudau XML загружен
❌ Удалено из файла (не в Розетке, кроме Мойдодыр/Dusel): {removed_missing}
⚠ Удалено как невалидных для MAUDAU: {removed_invalid}
🧩 Исключено (категория еще не заведена в MAUDAU): {removed_unknown_target_category}
❓ Офферов с категорией вне merchant_categories (обнаружено): {unresolved_target_category}
🗂 Переназначено категорий: {changed_category}
🏷 Добавлено/обновлено типовых параметров: {changed_params}
🆔 Скорректировано дублирующихся offer id: {deduped_ids}
💲 Обновлено цен: {changed_price}
🔁 Обновлено старых цен и наличия: {changed_other}
🧰 Режим Rozetka: {rozetka_mode}
🧰 Режим AquaFavorit: {base_mode}
📦 Отправляем на MAUDAU товаров: {kept}
📐 Размер итогового файла: {size_mb:.2f} MB
===== ГОТОВО ✅ ====="""

        if stale_alerts:
            report += "\n⚠ Источники недоступны более 72ч:\n" + "\n".join(stale_alerts)

        print(report)
        send_telegram(report)
        return 0
    except Exception as exc:
        error_msg = f"""===== СТАРТ =====
▶ Загрузка: Розетка XML
⚠ Ошибка: {exc}
===== ОШИБКА ❌ ====="""
        print(error_msg, file=sys.stderr)
        send_telegram(error_msg)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
